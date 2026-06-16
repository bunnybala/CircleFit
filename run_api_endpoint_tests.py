import time
import random
import string
import requests
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BASE_URL = "http://localhost:8081"
REPORT_PATH = r"d:\App\CircleFit\CircleFit_API_Test_Report.xlsx"

results = []

def log_result(tc_id, module, endpoint, test_case, payload_desc, expected, status, remarks=""):
    results.append((tc_id, module, endpoint, test_case, payload_desc, expected, status, remarks))
    icon = "[OK]" if status == "PASS" else "[FAIL]"
    print(f"  {icon} TC-{tc_id:02d} [{status}] {test_case} ({remarks})")

def generate_random_string(length=6):
    letters = string.ascii_lowercase
    return ''.join(random.choice(letters) for i in range(length))

def run_tests():
    print("Starting REST API Endpoint tests targeting:", BASE_URL)
    
    # Generate unique credentials for the test run
    rand = generate_random_string()
    username = f"api_user_{rand}"
    email = f"api_{rand}@circlefit.com"
    password = "password123"
    
    token = None
    group_id = None
    invite_code = None
    
    # ── 1. REGISTER ──────────────────────────────────────────
    tc_id = 1
    register_payload = {
        "username": username,
        "email": email,
        "password": password
    }
    try:
        start = time.time()
        res = requests.post(f"{BASE_URL}/api/auth/register", json=register_payload)
        duration = f"{time.time() - start:.3f}s"
        if res.status_code == 200:
            log_result(tc_id, "Auth", "POST /api/auth/register", "Register a new user successfully", 
                       f"POST {register_payload}", "Status 200 OK", "PASS", f"Success in {duration}")
        else:
            log_result(tc_id, "Auth", "POST /api/auth/register", "Register a new user successfully", 
                       f"POST {register_payload}", "Status 200 OK", "FAIL", f"Status {res.status_code}: {res.text} ({duration})")
    except Exception as e:
        log_result(tc_id, "Auth", "POST /api/auth/register", "Register a new user successfully", 
                   f"POST {register_payload}", "Status 200 OK", "FAIL", f"Exception: {str(e)}")

    # ── 2. REGISTER DUPLICATE USERNAME ──────────────────────
    tc_id += 1
    dup_uname_payload = {
        "username": username,
        "email": f"diff_{rand}@circlefit.com",
        "password": password
    }
    try:
        start = time.time()
        res = requests.post(f"{BASE_URL}/api/auth/register", json=dup_uname_payload)
        duration = f"{time.time() - start:.3f}s"
        if res.status_code == 500 or res.status_code == 400:
            log_result(tc_id, "Auth", "POST /api/auth/register", "Reject duplicate username", 
                       f"POST {dup_uname_payload}", "Error Status (400/500)", "PASS", f"Rejected username in {duration}")
        else:
            log_result(tc_id, "Auth", "POST /api/auth/register", "Reject duplicate username", 
                       f"POST {dup_uname_payload}", "Error Status (400/500)", "FAIL", f"Status {res.status_code} ({duration})")
    except Exception as e:
        log_result(tc_id, "Auth", "POST /api/auth/register", "Reject duplicate username", 
                   f"POST {dup_uname_payload}", "Error Status", "FAIL", str(e))

    # ── 3. REGISTER DUPLICATE EMAIL ─────────────────────────
    tc_id += 1
    dup_email_payload = {
        "username": f"diff_uname_{rand}",
        "email": email,
        "password": password
    }
    try:
        start = time.time()
        res = requests.post(f"{BASE_URL}/api/auth/register", json=dup_email_payload)
        duration = f"{time.time() - start:.3f}s"
        if res.status_code == 500 or res.status_code == 400:
            log_result(tc_id, "Auth", "POST /api/auth/register", "Reject duplicate email", 
                       f"POST {dup_email_payload}", "Error Status (400/500)", "PASS", f"Rejected email in {duration}")
        else:
            log_result(tc_id, "Auth", "POST /api/auth/register", "Reject duplicate email", 
                       f"POST {dup_email_payload}", "Error Status (400/500)", "FAIL", f"Status {res.status_code} ({duration})")
    except Exception as e:
        log_result(tc_id, "Auth", "POST /api/auth/register", "Reject duplicate email", 
                   f"POST {dup_email_payload}", "Error Status", "FAIL", str(e))

    # ── 4. LOGIN SUCCESS ─────────────────────────────────────
    tc_id += 1
    login_payload = {
        "email": email,
        "password": password
    }
    try:
        start = time.time()
        res = requests.post(f"{BASE_URL}/api/auth/login", json=login_payload)
        duration = f"{time.time() - start:.3f}s"
        if res.status_code == 200:
            body = res.json()
            token = body.get("token")
            log_result(tc_id, "Auth", "POST /api/auth/login", "Login with valid credentials", 
                       f"POST {login_payload}", "200 OK with token", "PASS", f"Received JWT in {duration}")
        else:
            log_result(tc_id, "Auth", "POST /api/auth/login", "Login with valid credentials", 
                       f"POST {login_payload}", "200 OK with token", "FAIL", f"Status {res.status_code}: {res.text}")
    except Exception as e:
        log_result(tc_id, "Auth", "POST /api/auth/login", "Login with valid credentials", 
                   f"POST {login_payload}", "200 OK with token", "FAIL", str(e))

    # ── 5. LOGIN INVALID CREDENTIALS ─────────────────────────
    tc_id += 1
    invalid_login_payload = {
        "email": email,
        "password": "wrong_password"
    }
    try:
        start = time.time()
        res = requests.post(f"{BASE_URL}/api/auth/login", json=invalid_login_payload)
        duration = f"{time.time() - start:.3f}s"
        if res.status_code in [401, 403, 500]:
            log_result(tc_id, "Auth", "POST /api/auth/login", "Login fails with incorrect password", 
                       f"POST {invalid_login_payload}", "401/403/500 Authentication Failure", "PASS", f"Access denied in {duration}")
        else:
            log_result(tc_id, "Auth", "POST /api/auth/login", "Login fails with incorrect password", 
                       f"POST {invalid_login_payload}", "401/403/500 Authentication Failure", "FAIL", f"Status {res.status_code} ({duration})")
    except Exception as e:
        log_result(tc_id, "Auth", "POST /api/auth/login", "Login fails with incorrect password", 
                   f"POST {invalid_login_payload}", "401/403 Unauthorized", "FAIL", str(e))

    # All subsequent requests need authorization headers
    headers = {"Authorization": f"Bearer {token}"} if token else {}

    # ── 6. GET PROFILE ────────────────────────────────────────
    tc_id += 1
    try:
        start = time.time()
        res = requests.get(f"{BASE_URL}/api/profile", headers=headers)
        duration = f"{time.time() - start:.3f}s"
        if res.status_code == 200:
            log_result(tc_id, "Profile", "GET /api/profile", "Retrieve profile data", 
                       "GET request with JWT", "200 OK containing user info", "PASS", f"Profile loaded in {duration}")
        else:
            log_result(tc_id, "Profile", "GET /api/profile", "Retrieve profile data", 
                       "GET request with JWT", "200 OK containing user info", "FAIL", f"Status {res.status_code} ({duration})")
    except Exception as e:
        log_result(tc_id, "Profile", "GET /api/profile", "Retrieve profile data", 
                   "GET request with JWT", "200 OK", "FAIL", str(e))

    # ── 7. PUT PROFILE UPDATE ─────────────────────────────────
    tc_id += 1
    profile_payload = {
        "name": "Live Tester",
        "age": 28,
        "height": 180.0,
        "weight": 78.5,
        "gender": "Male",
        "fitnessGoal": "Build Muscle",
        "dailyCalorieGoal": 2500
    }
    try:
        start = time.time()
        res = requests.put(f"{BASE_URL}/api/profile", json=profile_payload, headers=headers)
        duration = f"{time.time() - start:.3f}s"
        if res.status_code == 200:
            body = res.json()
            if body.get("name") == "Live Tester" and body.get("age") == 28:
                log_result(tc_id, "Profile", "PUT /api/profile", "Update profile details", 
                           f"PUT {profile_payload}", "200 OK with updated profile", "PASS", f"Updated in {duration}")
            else:
                log_result(tc_id, "Profile", "PUT /api/profile", "Update profile details", 
                           f"PUT {profile_payload}", "200 OK with updated profile", "FAIL", f"Updated but mismatch: {res.text}")
        else:
            log_result(tc_id, "Profile", "PUT /api/profile", "Update profile details", 
                       f"PUT {profile_payload}", "200 OK with updated profile", "FAIL", f"Status {res.status_code} ({duration})")
    except Exception as e:
        log_result(tc_id, "Profile", "PUT /api/profile", "Update profile details", 
                   f"PUT {profile_payload}", "200 OK", "FAIL", str(e))

    # ── 8. SYNC STEPS ─────────────────────────────────────────
    tc_id += 1
    today_str = datetime.now().strftime("%Y-%m-%d")
    steps_payload = [{
        "date": today_str,
        "steps": 7500,
        "distance": 5.4,
        "calories": 320.0
    }]
    try:
        start = time.time()
        res = requests.post(f"{BASE_URL}/api/steps/sync", json=steps_payload, headers=headers)
        duration = f"{time.time() - start:.3f}s"
        if res.status_code == 200:
            log_result(tc_id, "Steps", "POST /api/steps/sync", "Sync today's steps", 
                       f"POST {steps_payload}", "200 OK", "PASS", f"Synced in {duration}")
        else:
            log_result(tc_id, "Steps", "POST /api/steps/sync", "Sync today's steps", 
                       f"POST {steps_payload}", "200 OK", "FAIL", f"Status {res.status_code} ({duration})")
    except Exception as e:
        log_result(tc_id, "Steps", "POST /api/steps/sync", "Sync today's steps", 
                   f"POST {steps_payload}", "200 OK", "FAIL", str(e))

    # ── 9. GET WEEKLY STEPS ───────────────────────────────────
    tc_id += 1
    try:
        start = time.time()
        res = requests.get(f"{BASE_URL}/api/steps/weekly", headers=headers)
        duration = f"{time.time() - start:.3f}s"
        if res.status_code == 200:
            body = res.json()
            # We expect a list containing our sync record
            has_today = any(d.get("date") == today_str for d in body)
            log_result(tc_id, "Steps", "GET /api/steps/weekly", "Retrieve weekly step history", 
                       "GET weekly stats", "200 OK with daily stats list", "PASS" if has_today else "FAIL", 
                       f"Verified list in {duration}")
        else:
            log_result(tc_id, "Steps", "GET /api/steps/weekly", "Retrieve weekly step history", 
                       "GET weekly stats", "200 OK with daily stats list", "FAIL", f"Status {res.status_code} ({duration})")
    except Exception as e:
        log_result(tc_id, "Steps", "GET /api/steps/weekly", "Retrieve weekly step history", 
                   "GET weekly stats", "200 OK", "FAIL", str(e))

    # ── 10. CREATE GROUP ──────────────────────────────────────
    tc_id += 1
    group_payload = {
        "name": f"Team Run {rand}",
        "description": "API Test Group"
    }
    try:
        start = time.time()
        res = requests.post(f"{BASE_URL}/api/groups", json=group_payload, headers=headers)
        duration = f"{time.time() - start:.3f}s"
        if res.status_code == 200:
            body = res.json()
            group_id = body.get("id")
            invite_code = body.get("inviteCode")
            log_result(tc_id, "Groups", "POST /api/groups", "Create a fitness group", 
                       f"POST {group_payload}", "200 OK with Group details & inviteCode", "PASS", f"Group created: ID={group_id} ({duration})")
        else:
            log_result(tc_id, "Groups", "POST /api/groups", "Create a fitness group", 
                       f"POST {group_payload}", "200 OK", "FAIL", f"Status {res.status_code} ({duration})")
    except Exception as e:
        log_result(tc_id, "Groups", "POST /api/groups", "Create a fitness group", 
                   f"POST {group_payload}", "200 OK", "FAIL", str(e))

    # ── 11. GET MY GROUPS ─────────────────────────────────────
    tc_id += 1
    try:
        start = time.time()
        res = requests.get(f"{BASE_URL}/api/groups/my", headers=headers)
        duration = f"{time.time() - start:.3f}s"
        if res.status_code == 200:
            body = res.json()
            is_member = any(g.get("id") == group_id for g in body)
            log_result(tc_id, "Groups", "GET /api/groups/my", "Retrieve my groups list", 
                       "GET /api/groups/my", "200 OK containing list of user's groups", "PASS" if is_member else "FAIL", 
                       f"Found group membership in {duration}")
        else:
            log_result(tc_id, "Groups", "GET /api/groups/my", "Retrieve my groups list", 
                       "GET /api/groups/my", "200 OK", "FAIL", f"Status {res.status_code} ({duration})")
    except Exception as e:
        log_result(tc_id, "Groups", "GET /api/groups/my", "Retrieve my groups list", 
                   "GET /api/groups/my", "200 OK", "FAIL", str(e))

    # ── 12. GET LEADERBOARD ───────────────────────────────────
    tc_id += 1
    if group_id:
        try:
            start = time.time()
            res = requests.get(f"{BASE_URL}/api/groups/{group_id}/leaderboard", headers=headers)
            duration = f"{time.time() - start:.3f}s"
            if res.status_code == 200:
                body = res.json()
                has_creator = any(e.get("username") == username for e in body)
                log_result(tc_id, "Groups", "GET /api/groups/{id}/leaderboard", "Fetch group leaderboard", 
                           f"GET leaderboard for group {group_id}", "200 OK containing rankings", "PASS" if has_creator else "FAIL", 
                           f"Leaderboard parsed in {duration}")
            else:
                log_result(tc_id, "Groups", "GET /api/groups/{id}/leaderboard", "Fetch group leaderboard", 
                           f"GET leaderboard for group {group_id}", "200 OK", "FAIL", f"Status {res.status_code} ({duration})")
        except Exception as e:
            log_result(tc_id, "Groups", "GET /api/groups/{id}/leaderboard", "Fetch group leaderboard", 
                       f"GET leaderboard for group {group_id}", "200 OK", "FAIL", str(e))
    else:
        log_result(tc_id, "Groups", "GET /api/groups/{id}/leaderboard", "Fetch group leaderboard", 
                   "GET leaderboard", "200 OK", "FAIL", "Skipped: group_id not generated")

    # ── 13. GET CHAT HISTORY ──────────────────────────────────
    tc_id += 1
    if group_id:
        try:
            start = time.time()
            res = requests.get(f"{BASE_URL}/api/groups/{group_id}/chat/history", headers=headers)
            duration = f"{time.time() - start:.3f}s"
            if res.status_code == 200:
                log_result(tc_id, "Chat", "GET /api/groups/{id}/chat/history", "Retrieve group chat history", 
                           f"GET chat history for group {group_id}", "200 OK containing list", "PASS", f"Retrieved in {duration}")
            else:
                log_result(tc_id, "Chat", "GET /api/groups/{id}/chat/history", "Retrieve group chat history", 
                           f"GET chat history for group {group_id}", "200 OK", "FAIL", f"Status {res.status_code} ({duration})")
        except Exception as e:
            log_result(tc_id, "Chat", "GET /api/groups/{id}/chat/history", "Retrieve group chat history", 
                       f"GET chat history for group {group_id}", "200 OK", "FAIL", str(e))
    else:
        log_result(tc_id, "Chat", "GET /api/groups/{id}/chat/history", "Retrieve group chat history", 
                   "GET chat history", "200 OK", "FAIL", "Skipped: group_id not generated")

def generate_excel_report(filepath):
    wb = Workbook()
    ws = wb.active
    ws.title = "REST API Test Report"

    # Styles
    title_fill = PatternFill(start_color="800080", end_color="800080", fill_type="solid") # Deep Purple
    header_fill = PatternFill(start_color="4B0082", end_color="4B0082", fill_type="solid") # Indigo
    
    pass_font = Font(name="Calibri", bold=True, color="276221")
    pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    
    fail_font = Font(name="Calibri", bold=True, color="9C0006")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )
    wrap = Alignment(vertical="center", wrap_text=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Column widths
    widths = [6, 12, 35, 40, 40, 30, 12, 40]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    # Title
    ws.merge_cells("A1:H1")
    ws["A1"] = "CircleFit — REST API Integration Test Report"
    ws["A1"].font = Font(name="Calibri", bold=True, color="FFFFFF", size=14)
    ws["A1"].fill = title_fill
    ws["A1"].alignment = center
    ws.row_dimensions[1].height = 30

    # Date
    ws.merge_cells("A2:H2")
    ws["A2"] = f"Generated: {datetime.now().strftime('%d %B %Y  %H:%M')}  |  Client: Python requests library"
    ws["A2"].alignment = wrap
    ws.row_dimensions[2].height = 20

    # Summary
    total = len(results)
    passed = sum(1 for r in results if r[6] == "PASS")
    failed = total - passed
    rate = (passed * 100 // total) if total > 0 else 0
    
    ws.merge_cells("A4:H4")
    ws["A4"] = f"Total Endpoints Tested: {total}    PASS: {passed}    FAIL: {failed}    Pass Rate: {rate}%"
    ws["A4"].font = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
    ws["A4"].fill = title_fill
    ws["A4"].alignment = center
    ws.row_dimensions[4].height = 24

    # Headers
    headers = ["#", "Module", "Endpoint", "Test Case", "Request Payload", "Expected Result", "Status", "Response / Remarks"]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=6, column=i, value=h)
        cell.font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
    ws.row_dimensions[6].height = 36

    # Data rows
    for idx, r in enumerate(results):
        row_idx = idx + 7
        for col_idx, val in enumerate(r):
            cell = ws.cell(row=row_idx, column=col_idx + 1, value=str(val))
            cell.border = border
            cell.alignment = wrap
            if col_idx == 0:
                cell.alignment = center
            if col_idx == 6:  # Status column
                if val == "PASS":
                    cell.font = pass_font
                    cell.fill = pass_fill
                else:
                    cell.font = fail_font
                    cell.fill = fail_fill
                cell.alignment = center
        ws.row_dimensions[row_idx].height = 40

    wb.save(filepath)
    print(f"\n[OK] API Test Report saved to: {filepath}")
    print(f"     Total: {total}  PASS: {passed}  FAIL: {failed}  Rate: {rate}%")

def main():
    run_tests()
    generate_excel_report(REPORT_PATH)

if __name__ == "__main__":
    main()
