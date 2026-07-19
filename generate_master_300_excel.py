import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def main():
    wb = openpyxl.Workbook()
    
    # Color palette and typography
    font_family = "Segoe UI"
    
    font_title = Font(name=font_family, size=16, bold=True, color="FFFFFF")
    font_section = Font(name=font_family, size=11, bold=True, color="3F3D56")
    font_header = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    font_body = Font(name=font_family, size=10, color="333333")
    font_bold = Font(name=font_family, size=10, bold=True, color="333333")
    
    font_card_num = Font(name=font_family, size=16, bold=True, color="2B2D42")
    font_card_lbl = Font(name=font_family, size=9, bold=True, color="8D99AE")
    
    fill_purple_header = PatternFill(start_color="6C63FF", end_color="6C63FF", fill_type="solid")
    fill_light_purple = PatternFill(start_color="F4F3FF", end_color="F4F3FF", fill_type="solid")
    fill_card = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
    fill_total_row = PatternFill(start_color="EAE8FF", end_color="EAE8FF", fill_type="solid")
    
    thin_border_side = Side(style="thin", color="D1D5DB")
    border_cell = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    border_header = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=Side(style="medium", color="4B44D4"))
    border_card = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    border_total = Border(top=Side(style="double", color="3F3D56"), bottom=Side(style="double", color="3F3D56"))
    
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # ─── 1. SUMMARY DASHBOARD SHEET ──────────────────────────────────────────
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary.views.sheetView[0].showGridLines = True
    
    # Title Block
    ws_summary.merge_cells("A1:G2")
    title_cell = ws_summary["A1"]
    title_cell.value = "CircleFit — Master QA Test Plan Dashboard"
    title_cell.font = font_title
    title_cell.fill = fill_purple_header
    title_cell.alignment = align_center
    
    # Subtitle
    ws_summary.merge_cells("A3:G3")
    subtitle_cell = ws_summary["A3"]
    subtitle_cell.value = "Quality assurance master repository consisting of 300 test cases across Mobile, Web, and API modules."
    subtitle_cell.font = Font(name=font_family, size=10, italic=True, color="555555")
    subtitle_cell.alignment = align_left
    
    # KPI 1: Total Appium Cases
    ws_summary["A4"] = "APPIUM CASES"
    ws_summary["A4"].font = font_card_lbl
    ws_summary["A4"].alignment = align_center
    ws_summary["A4"].fill = fill_card
    ws_summary["A4"].border = border_card
    
    ws_summary.merge_cells("A5:A6")
    ws_summary["A5"] = "=B11"
    ws_summary["A5"].font = font_card_num
    ws_summary["A5"].alignment = align_center
    ws_summary["A5"].fill = fill_card
    ws_summary["A5"].border = border_card
    
    # KPI 2: Total Selenium Cases
    ws_summary["B4"] = "SELENIUM CASES"
    ws_summary["B4"].font = font_card_lbl
    ws_summary["B4"].alignment = align_center
    ws_summary["B4"].fill = fill_card
    ws_summary["B4"].border = border_card
    
    ws_summary.merge_cells("B5:B6")
    ws_summary["B5"] = "=B12"
    ws_summary["B5"].font = font_card_num
    ws_summary["B5"].alignment = align_center
    ws_summary["B5"].fill = fill_card
    ws_summary["B5"].border = border_card
    
    # KPI 3: Total API Cases
    ws_summary["C4"] = "API CASES"
    ws_summary["C4"].font = font_card_lbl
    ws_summary["C4"].alignment = align_center
    ws_summary["C4"].fill = fill_card
    ws_summary["C4"].border = border_card
    
    ws_summary.merge_cells("C5:C6")
    ws_summary["C5"] = "=B13"
    ws_summary["C5"].font = font_card_num
    ws_summary["C5"].alignment = align_center
    ws_summary["C5"].fill = fill_card
    ws_summary["C5"].border = border_card
    
    # KPI 4: Grand Total
    ws_summary["D4"] = "TOTAL TEST SUITE"
    ws_summary["D4"].font = Font(name=font_family, size=9, bold=True, color="FFFFFF")
    ws_summary["D4"].alignment = align_center
    ws_summary["D4"].fill = fill_purple_header
    ws_summary["D4"].border = border_card
    
    ws_summary.merge_cells("D5:G6")
    ws_summary["D5"] = "=B14"
    ws_summary["D5"].font = Font(name=font_family, size=20, bold=True, color="FFFFFF")
    ws_summary["D5"].alignment = align_center
    ws_summary["D5"].fill = fill_purple_header
    ws_summary["D5"].border = border_card
    
    # Section Label
    ws_summary["A8"] = "TEST SUITE COMPONENT BREAKDOWN"
    ws_summary["A8"].font = font_section
    
    # Headers for breakdown table
    breakdown_headers = ["Test Category", "Total Cases", "Scope", "Reference Sheet", "Primary Tech Stack", "Description"]
    ws_summary.merge_cells("F10:G10")
    
    for col_idx, text in enumerate(breakdown_headers, 1):
        cell = ws_summary.cell(row=10, column=col_idx, value=text)
        cell.font = font_header
        cell.fill = fill_purple_header
        cell.alignment = align_center
        cell.border = border_header
        
    # Categories rows
    # Row 11: Appium
    ws_summary["A11"] = "Appium Mobile"
    ws_summary["B11"] = "=COUNTA('Appium Mobile'!A7:A106)"
    ws_summary["C11"] = "Mobile UI, Sensors, Background isolates, Local DB caching"
    ws_summary["D11"] = "Appium Mobile"
    ws_summary["E11"] = "Flutter / Java / Python / Appium"
    ws_summary.merge_cells("F11:G11")
    ws_summary["F11"] = "Automated integration and UI test suite targeting Android/iOS packages."
    
    # Row 12: Selenium
    ws_summary["A12"] = "Selenium Web"
    ws_summary["B12"] = "=COUNTA('Selenium Web'!A7:A106)"
    ws_summary["C12"] = "Responsive views, State interceptors, SVG charts, Modals"
    ws_summary["D12"] = "Selenium Web"
    ws_summary["E12"] = "React / Vite / Python / Selenium"
    ws_summary.merge_cells("F12:G12")
    ws_summary["F12"] = "Automated browser visual regression and cross-browser alignment suite."
    
    # Row 13: API
    ws_summary["A13"] = "REST & WebSocket API"
    ws_summary["B13"] = "=COUNTA('REST & WebSocket API'!A7:A106)"
    ws_summary["C13"] = "JWT Auth, Rate limits, WebSocket STOMP brokers, CORS rules"
    ws_summary["D13"] = "REST & WebSocket API"
    ws_summary["E13"] = "Spring Boot / Java / STOMP WebSockets"
    ws_summary.merge_cells("F13:G13")
    ws_summary["F13"] = "Backend integration endpoints payload validation, security, and stress tests."
    
    # Row 14: Total
    ws_summary["A14"] = "Total Suite Cases"
    ws_summary["B14"] = "=SUM(B11:B13)"
    ws_summary["C14"] = "Full coverage of CircleFit core endpoints and client experiences."
    ws_summary["D14"] = "All Sheets"
    ws_summary["E14"] = "Python request runners / Drivers"
    ws_summary.merge_cells("F14:G14")
    ws_summary["F14"] = "Consolidated QA test coverage metrics."
    
    # Style Breakdown Table
    for row in range(11, 15):
        fill = fill_total_row if row == 14 else (fill_light_purple if row % 2 == 0 else PatternFill(fill_type=None))
        font = font_bold if row == 14 else font_body
        border = border_total if row == 14 else border_cell
        
        ws_summary.cell(row=row, column=1).font = font
        ws_summary.cell(row=row, column=1).fill = fill
        ws_summary.cell(row=row, column=1).alignment = align_left
        ws_summary.cell(row=row, column=1).border = border
        
        ws_summary.cell(row=row, column=2).font = font
        ws_summary.cell(row=row, column=2).fill = fill
        ws_summary.cell(row=row, column=2).alignment = align_center
        ws_summary.cell(row=row, column=2).border = border
        
        for col in range(3, 6):
            cell = ws_summary.cell(row=row, column=col)
            cell.font = font
            cell.fill = fill
            cell.alignment = align_left
            cell.border = border
            
        # Merge cell boundaries require formatting both cells in merge
        cell_f = ws_summary.cell(row=row, column=6)
        cell_g = ws_summary.cell(row=row, column=7)
        cell_f.font = font; cell_f.fill = fill; cell_f.alignment = align_left; cell_f.border = border
        cell_g.font = font; cell_g.fill = fill; cell_g.alignment = align_left; cell_g.border = border
        
    ws_summary.row_dimensions[1].height = 20
    ws_summary.row_dimensions[2].height = 20
    ws_summary.row_dimensions[4].height = 18
    ws_summary.row_dimensions[5].height = 18
    ws_summary.row_dimensions[6].height = 18
    ws_summary.row_dimensions[10].height = 28
    for r in range(11, 15):
        ws_summary.row_dimensions[r].height = 24
        
    ws_summary.column_dimensions['A'].width = 22
    ws_summary.column_dimensions['B'].width = 14
    ws_summary.column_dimensions['C'].width = 45
    ws_summary.column_dimensions['D'].width = 22
    ws_summary.column_dimensions['E'].width = 28
    ws_summary.column_dimensions['F'].width = 35
    ws_summary.column_dimensions['G'].width = 35
    
    # ─── HELPER FOR CREATING DATA SHEETS ──────────────────────────────────────
    def populate_data_sheet(sheet_title, headers, data):
        ws = wb.create_sheet(title=sheet_title)
        ws.views.sheetView[0].showGridLines = True
        
        # Header Banner
        ws.merge_cells("A1:F1")
        banner = ws["A1"]
        banner.value = f"CircleFit — {sheet_title} Test Suite"
        banner.font = font_title
        banner.fill = fill_purple_header
        banner.alignment = align_center
        ws.row_dimensions[1].height = 28
        
        # Metadata Block
        ws.merge_cells("A2:F2")
        meta = ws["A2"]
        meta.value = f"Automated Master Log  |  Total Cases: {len(data)}  |  Category: {sheet_title}"
        meta.font = Font(name=font_family, size=10, italic=True, color="555555")
        meta.alignment = align_left
        ws.row_dimensions[2].height = 18
        
        # Spacer row
        ws.row_dimensions[3].height = 10
        
        # Summary row
        ws.merge_cells("A4:F4")
        summary_cell = ws["A4"]
        summary_cell.value = f"Component Cases count: =COUNTA(A7:A{len(data)+6})"
        summary_cell.font = font_bold
        summary_cell.alignment = align_left
        summary_cell.fill = fill_light_purple
        ws.row_dimensions[4].height = 24
        
        # Spacer
        ws.row_dimensions[5].height = 10
        
        # Headers row
        for col_idx, text in enumerate(headers, 1):
            cell = ws.cell(row=6, column=col_idx, value=text)
            cell.font = font_header
            cell.fill = fill_purple_header
            cell.alignment = align_center
            cell.border = border_header
        ws.row_dimensions[6].height = 30
        
        # Data rows
        for idx, row_data in enumerate(data, 7):
            ws.row_dimensions[idx].height = 36
            for col_idx, val in enumerate(row_data, 1):
                cell = ws.cell(row=idx, column=col_idx, value=val)
                cell.font = font_body
                cell.border = border_cell
                
                # Alignments
                if col_idx in [1]:
                    cell.alignment = align_center
                else:
                    cell.alignment = align_left
                    
        # Column widths
        widths = [10, 16, 18, 45, 45, 55]
        for c_idx, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(c_idx)].width = w

    # ─── 2. APPIUM MOBILE DATA (100 Cases) ──────────────────────────────────
    appium_headers = ["ID", "Module", "Screen", "Scenario / Description", "Preconditions & Inputs", "Expected Result"]
    appium_data = [
        ("APP-001", "Auth", "Splash", "Verify app launches without crash", "APK installed, open application", "App loads, renders first splash screen"),
        ("APP-002", "Auth", "Login", "Verify Email field exists", "Launch app, land on login screen", "Email input field is visible"),
        ("APP-003", "Auth", "Login", "Verify Password field exists", "Launch app, land on login screen", "Password input field is visible"),
        ("APP-004", "Auth", "Login", "Verify Login button exists", "Launch app, land on login screen", "Login button is visible and active"),
        ("APP-005", "Auth", "Login", "Verify Register link redirects to Signup screen", "Tap 'Don't have an account? Register'", "Navigates user to SignUp screen"),
        ("APP-006", "Auth", "Login", "Blank password submission validation", "Enter email, leave password blank, tap Login", "Email validated, error 'Password is required'"),
        ("APP-007", "Auth", "Login", "Blank email submission validation", "Leave email blank, enter password, tap Login", "Error 'Email is required' is shown"),
        ("APP-008", "Auth", "Login", "Malformed email syntax validation", "Enter 'invalidemail', tap Login", "Error 'Invalid email format' displayed"),
        ("APP-009", "Auth", "Login", "Valid login credentials authorization", "Enter 'test@circlefit.com' & 'test123', login", "Logs in, fetches JWT, goes to Dashboard"),
        ("APP-010", "Auth", "Login", "Invalid credentials validation error", "Enter wrong credentials, tap Login", "Error message shown: 'Invalid email or password'"),
        ("APP-011", "Auth", "Register", "Username field is visible", "Navigate to SignUp screen", "Username input is rendered"),
        ("APP-012", "Auth", "Register", "Email field is visible", "Navigate to SignUp screen", "Email input is rendered"),
        ("APP-013", "Auth", "Register", "Password field is visible", "Navigate to SignUp screen", "Password input is rendered"),
        ("APP-014", "Auth", "Register", "Register button validation", "Navigate to SignUp screen", "Register button is disabled if fields empty"),
        ("APP-015", "Auth", "Splash", "Automatic login with valid stored JWT", "Launch app while valid token exists", "Splash screen redirects straight to Dashboard"),
        ("APP-016", "Dashboard", "Dashboard", "Step count circular progress gauge render", "Authenticated, land on Dashboard", "Gauge widget renders current steps visually"),
        ("APP-017", "Dashboard", "Dashboard", "Steps increment from physical pedometer sensor", "Increase steps on emulator/device sensor", "Circular progress gauge increments in real-time"),
        ("APP-018", "Dashboard", "Dashboard", "Real-time calculations of distance", "Walk 1000 steps", "Distance increases by 0.76 km (steps * 0.000762)"),
        ("APP-019", "Dashboard", "Dashboard", "Real-time calculations of active calories", "Walk 1000 steps", "Calorie display increments by 40 kcal (steps * 0.04)"),
        ("APP-020", "Dashboard", "Dashboard", "Midnight steps count rollover reset", "Run app at 23:59:59, wait for midnight", "Daily steps reset to 0; total logs to history DB"),
        ("APP-021", "Dashboard", "Dashboard", "Device reboot baseline recovery", "Pedometer baseline reads 5000, reboot device", "Baseline reset to 200, steps increment from 0"),
        ("APP-022", "Dashboard", "Dashboard", "Manual pull-to-refresh steps sync", "Pull down dashboard screen list", "Spins loader, uploads local cached steps to DB"),
        ("APP-023", "Dashboard", "Dashboard", "Automatic background steps sync timer", "Keep app running in background for 15m", "Steps sync automatically via background task"),
        ("APP-024", "Dashboard", "Dashboard", "Steps sync with 0 values handled", "Sync steps with 0 new steps walked", "Sync request skipped or returns safely, no error"),
        ("APP-025", "Dashboard", "Dashboard", "Extreme step counts upload boundary check", "Walk 999,999 steps, trigger sync", "App safely truncates or warns of invalid range"),
        ("APP-026", "Groups", "Browser", "List communities joined by user", "Tap Groups tab", "Displays joined groups list with titles and goals"),
        ("APP-027", "Groups", "Browser", "Create new fitness group", "Tap 'Create Group', enter details, submit", "Group saved, creator assigned admin badge"),
        ("APP-028", "Groups", "Browser", "Generate 6-char alphanumeric invite code", "Create a new group", "Displays unique 6-char code (e.g., AB12CD)"),
        ("APP-029", "Groups", "Browser", "Join group using valid invite code", "Enter code 'AB12CD', click Join", "Success toast shown; group added to user's list"),
        ("APP-030", "Groups", "Browser", "Prevent joining duplicate groups", "Enter code for group user already belongs to", "Toast error: 'Already a member of this group'"),
        ("APP-031", "Groups", "Detail", "Roster list of members rendering", "Tap on a group card", "Renders members list with names, avatars, steps"),
        ("APP-032", "Groups", "Detail", "Leaderboard ranking descending order", "View leaderboard section of group", "Rankings sorted with highest step member at top"),
        ("APP-033", "Groups", "Detail", "Leaderboard podium visualization", "View leaderboard podium", "First place sits highest, second/third adjusted"),
        ("APP-034", "Groups", "Detail", "Group challenges tab loads list", "Select Challenges tab in group details", "Challenges list with participants load correctly"),
        ("APP-035", "Groups", "Detail", "Create group challenge form", "Tap 'New Challenge' from challenges tab", "Navigates to challenge setup form UI"),
        ("APP-036", "Chat", "Chat", "STOMP WS endpoint handshake connection", "Tap Chat tab in group", "WebSocket connects successfully (status: Connected)"),
        ("APP-037", "Chat", "Chat", "Send messages over WebSocket connection", "Type message, click send", "Transmits JSON over STOMP stream; renders bubble"),
        ("APP-038", "Chat", "Chat", "Receive messages in real-time", "Receive socket message from member B", "Renders message instantly in chat timeline bubble"),
        ("APP-039", "Chat", "Chat", "Retrieve group message history from API", "Open chat thread with existing history", "Historical message bubbles render in chat stream"),
        ("APP-040", "Chat", "Chat", "STOMP client auto-reconnect on drops", "Disconnect network data, then reconnect", "Auto-reconnects, re-subscribes without data loss"),
        ("APP-041", "Chat", "Chat", "Special characters support in chat", "Send message containing emojis and symbols", "Message renders correctly in chat bubble"),
        ("APP-042", "Chat", "Chat", "Block blank message submissions", "Try sending empty text", "Send button disabled; no blank bubbles allowed"),
        ("APP-043", "Profile", "Profile", "User details populate on profile display", "Tap Profile tab", "Username, email, height, weight populate matching DB"),
        ("APP-044", "Profile", "Profile", "Modify height validation range constraints", "Enter height 40cm or 350cm, save", "Input blocks, shows height error boundaries"),
        ("APP-045", "Profile", "Profile", "Modify weight validation range constraints", "Enter weight 10kg or 600kg, save", "Input blocks, shows weight error boundaries"),
        ("APP-046", "Profile", "Profile", "Modify age validation range constraints", "Enter age 2 or 150, save", "Input blocks, shows age error boundaries"),
        ("APP-047", "Profile", "Profile", "Profile updates persist on database", "Edit height to 180cm, age to 28, save", "Saves locally, pushes API update, status 200"),
        ("APP-048", "Profile", "Profile", "Profile picture avatar initials default", "Display profile before uploading photo", "CircleAvatar renders initials of username"),
        ("APP-049", "Profile", "Profile", "Achievements dynamic badge unlocking", "Walk 10,000 steps daily", "'10K Step Master' badge turns active color"),
        ("APP-050", "Profile", "Profile", "Log out termination session redirects", "Tap Logout button in settings", "Clears local storage JWT, loads Login screen"),
        ("APP-051", "Nutrition", "Dashboard", "Hydration water intake quick-add click", "Tap +250ml quick-add button", "Hydration count updates; water wave animation rises"),
        ("APP-052", "Nutrition", "Dashboard", "Hydration daily limit cap (10 Liters)", "Tap +250ml repeatedly up to 10L", "Capped at 10,000ml to prevent overflow metrics"),
        ("APP-053", "Nutrition", "Dashboard", "Hydration reset daily log action", "Tap reset button, confirm modal", "Hydration volume resets back to 0ml"),
        ("APP-054", "Nutrition", "Dashboard", "Food search navigates and displays results", "Tap Search Food, type 'Oats'", "Lists product cards retrieved from Open Food Facts API"),
        ("APP-055", "Nutrition", "Dashboard", "Food preset scanner layout guide", "Tap Scan Food barcode preset", "Visual laser line guide renders, matches overlay"),
        ("APP-056", "Platform", "Global", "Orientation landscape layout adapting", "Set emulator to landscape mode", "Views wrap, scrollbars prevent overflow warnings"),
        ("APP-057", "Platform", "Global", "Android system font scale scaling", "Increase system font to 150% in settings", "Labels wrap or expand without layout clipping"),
        ("APP-058", "Platform", "Global", "Keyboard dismissal gesture check", "Tap on blank space outside text field", "Focus dismissed, keyboard slides down"),
        ("APP-059", "Platform", "Global", "Keyboard avoids covering submission buttons", "Tap text inputs on Login screen", "Action buttons scroll up, remain tapable"),
        ("APP-060", "Platform", "Global", "Step logging local caching when offline", "Walk 50 steps while network disconnected", "Writes steps to SQLite DB, shows 'Cached' tag"),
        ("APP-061", "Platform", "Global", "Memory warning triggers SQLite dump", "Trigger system low memory warning", "App writes RAM steps cache to SQLite DB safely"),
        ("APP-062", "Platform", "Global", "Physical activity permission denial flow", "Decline physical activity recognition prompt", "Renders rationale overlay explaining app benefits"),
        ("APP-063", "Platform", "Global", "Expired session token redirect interceptor", "Make API request with expired token", "401 code caught, token cleared, redirects to Login"),
        ("APP-064", "Platform", "Global", "SQL injection inputs sanitization", "Type ' OR '1'='1 in fields", "Treated as literal text, no backend crash"),
        ("APP-065", "Platform", "Global", "Android back-button pop validation", "Tap back button on home dashboard", "App does not pop to login, closes app or stays"),
        ("APP-066", "Auth", "Splash", "App upgrade notification dialog", "Launch app with outdated version", "Pop-up prompts user to update to latest build"),
        ("APP-067", "Auth", "Register", "Password mismatch validation", "Enter password 'pass123', confirm 'pass321'", "Error 'Passwords do not match' displayed"),
        ("APP-068", "Dashboard", "Dashboard", "Weekly step chart updates from DB history", "Load dashboard after sync", "Chart loads 7 bars representing daily step history"),
        ("APP-069", "Dashboard", "Dashboard", "Step gauge responsiveness", "Walk 10 steps, check gauge update", "Circular indicator shifts in under 500ms"),
        ("APP-070", "Dashboard", "Dashboard", "App minimized background step sync", "Minimize app, walk 50 steps, maximize", "Steps synced and loaded into gauge correctly"),
        ("APP-071", "Dashboard", "Dashboard", "Steps tracking service persistent notification", "Minimize app", "Notification drawer shows active background service"),
        ("APP-072", "Dashboard", "Dashboard", "Steps counter sync status icon", "Sync steps successfully", "Status icon changes from warning yellow to green"),
        ("APP-073", "Dashboard", "Dashboard", "Hydration logs persist across restarts", "Log 500ml water, restart app", "Relaunches displaying 500ml water logged"),
        ("APP-074", "Dashboard", "Dashboard", "Calorie balance needle gauge calculation", "Consume 2000 kcal, burn 500 kcal", "Pointer shifts right to show surplus calorie balance"),
        ("APP-075", "Dashboard", "Dashboard", "BMI color-coded card levels", "Profile BMI = 18.0 (Underweight)", "BMI status card turns blue (Underweight styling)"),
        ("APP-076", "Dashboard", "Dashboard", "BMI normal weight level color", "Profile BMI = 22.0 (Normal)", "BMI status card turns green (Normal styling)"),
        ("APP-077", "Dashboard", "Dashboard", "BMI overweight level color", "Profile BMI = 27.0 (Overweight)", "BMI status card turns yellow (Overweight styling)"),
        ("APP-078", "Dashboard", "Dashboard", "BMI obese level color", "Profile BMI = 32.0 (Obese)", "BMI status card turns red (Obese styling)"),
        ("APP-079", "Dashboard", "Dashboard", "Network connection offline toast notification", "Disconnect internet data", "Bottom toast alerts: 'No Internet. Running offline'"),
        ("APP-080", "Dashboard", "Dashboard", "Network connection back online toast", "Reconnect internet data", "Toast alerts: 'Internet restored. Syncing data...'"),
        ("APP-081", "Groups", "Detail", "User can leave joined community", "Tap 'Leave Group' in details", "Removed from group; roster steps updated"),
        ("APP-082", "Groups", "Detail", "Admin cannot leave group without transfer", "Tap 'Leave Group' as creator", "Warning: 'Assign new group admin before leaving'"),
        ("APP-083", "Groups", "Detail", "Empty leaderboard state handler", "Join a new group with 0 steps", "Leaderboard shows user list with 0 steps safely"),
        ("APP-084", "Groups", "Detail", "Group step goal progress bar", "Open group details", "Displays progress bar of group average vs step goal"),
        ("APP-085", "Groups", "Detail", "Member profile card popup dialog", "Tap on user avatar in group roster", "Dialog shows member's weekly steps history chart"),
        ("APP-086", "Groups", "Detail", "Alphanumeric code invite clipboard copy", "Tap copy icon next to invite code", "Code copied to clipboard; toast notification shown"),
        ("APP-087", "Chat", "Chat", "Scroll-to-bottom on new messages", "Send/receive message while scrolled up", "Chat window does not force scroll if user scrolling up"),
        ("APP-088", "Chat", "Chat", "Keyboard layout shifts chat input box", "Tap text input field", "Input bar shifts up above keyboard cleanly"),
        ("APP-089", "Chat", "Chat", "Long text message wraps inside bubble", "Send 200-word paragraph message", "Text wraps to fit screen width, bubble resizes"),
        ("APP-090", "Chat", "Chat", "Message timestamp visibility", "Long press on a chat bubble", "Renders exact date/time timestamp beneath bubble"),
        ("APP-091", "Profile", "Profile", "BMR calculation defaults on empty fields", "Clear weight/height, save", "BMR calculation falls back to 1800 kcal default"),
        ("APP-092", "Profile", "Profile", "Age boundaries reject decimals input", "Enter age 25.5, click save", "Input validated, restricts to integers only"),
        ("APP-093", "Profile", "Profile", "Height boundaries reject negative numbers", "Enter height -160cm, click save", "Form validation prevents positive boundary bypass"),
        ("APP-094", "Profile", "Profile", "Weight boundaries reject negative numbers", "Enter weight -70kg, click save", "Form validation prevents positive boundary bypass"),
        ("APP-095", "Nutrition", "Dashboard", "Add calories to daily logs updates needle", "Search oats, add 300 kcal", "consumed calories increases; surplus needle updates"),
        ("APP-096", "Nutrition", "Dashboard", "Food search empty keyword validation", "Tap search icon with empty text", "Search blocked; toast alert: 'Enter keyword to search'"),
        ("APP-097", "Nutrition", "Dashboard", "Scan preset Apple details validation", "Tap Apple scanning preset", "Resolves Apple details, calories, and macro values"),
        ("APP-098", "Platform", "Global", "Application state minimized lifecycle", "Minimize app, lock screen, unlock", "App restores foreground state without refresh errors"),
        ("APP-099", "Platform", "Global", "Navigation tab index persistent state", "Tap Profile, minimize, restore app", "App returns foreground maintaining Profile tab index"),
        ("APP-100", "Platform", "Global", "Double-tap back button exit flow", "Double tap Android back button rapidly", "Triggers clean app exit closing active processes")
    ]
    populate_data_sheet("Appium Mobile", appium_headers, appium_data)

    # ─── 3. SELENIUM WEB DATA (100 Cases) ────────────────────────────────────
    selenium_headers = ["ID", "Module", "Screen", "Scenario / Description", "Preconditions & Inputs", "Expected Result"]
    selenium_data = [
        ("SEL-001", "Auth", "Login", "Verify login card vertical alignment", "Open page in browser", "Login card is centered vertically on screen"),
        ("SEL-002", "Auth", "Login", "Blank form input submission warning", "Click login with blank fields", "Warning 'Please fill in all fields' is shown"),
        ("SEL-003", "Auth", "Login", "Browser HTML5 email verification", "Input 'user.com', click Login", "Browser validator blocks and prompts format error"),
        ("SEL-004", "Auth", "Login", "Token writing to localStorage", "Enter valid login credentials", "JWT token saved inside localStorage key"),
        ("SEL-005", "Auth", "Main", "Token deletion on logout", "Click Logout, confirm modal", "localStorage JWT token removed; session cleared"),
        ("SEL-006", "Platform", "Global", "Routing guard redirects direct URL access", "Navigate to /dashboard directly", "Intercepts, redirects back to /login layout"),
        ("SEL-007", "Platform", "Global", "Sidebar collapses to drawer on mobile", "Resize viewport to 600px", "Sidebar collapses to toggleable hamburger drawer"),
        ("SEL-008", "Platform", "Global", "Main content panel responsive padding", "Resize viewport from 1200px to 800px", "Padding adjusts dynamically; dashboard elements wrap"),
        ("SEL-009", "Platform", "Global", "Dashboard CSS grid wrapping", "View homepage cards", "Grid layouts wrap items smoothly to prevent overflows"),
        ("SEL-010", "Dashboard", "Dashboard", "Step progress circular SVG path render", "Open dashboard", "Radial SVG indicator matches daily step goals"),
        ("SEL-011", "Dashboard", "Dashboard", "Weekly step chart heights scale", "Load dashboard with step data", "Columns heights auto-scale based on daily step max"),
        ("SEL-012", "Dashboard", "Dashboard", "Hydration wave visual progress heights", "Update logged water quantity", "Water wave levels rise/fall based on goals percentage"),
        ("SEL-013", "Dashboard", "Dashboard", "Needle slider balance Surplus/Deficit", "Update consumed calories metrics", "Needle pointer offsets left/right dynamically"),
        ("SEL-014", "Dashboard", "Dashboard", "Premium Typography family verification", "Inspect page headers in DevTools", "Font-family parses Roboto/Segoe UI body fonts"),
        ("SEL-015", "Dashboard", "Dashboard", "Primary button background contrast standard", "Inspect purple button #6C63FF", "Contrast ratio exceeds 4.5:1 WCAG AA guidelines"),
        ("SEL-016", "Platform", "Global", "Text inputs consistent padding metrics", "Inspect focus inputs styles", "Standard margins, rounded borders, clear outlines"),
        ("SEL-017", "Platform", "Global", "Navigation sidebar links hover states", "Hover cursor over menu items", "Background color transitions smoothly on hover"),
        ("SEL-018", "Platform", "Global", "Dashboard skeleton loader placeholder", "Reload dashboard on slow throttle network", "Shimmer blocks load before actual cards display"),
        ("SEL-019", "Platform", "Global", "Submit buttons loading state behaviors", "Click Save profile settings", "Button labels dim, loader spins, clicks are blocked"),
        ("SEL-020", "Platform", "Global", "Field validation borders transition", "Submit registration with error", "Field border outlines turn red; error labels appear"),
        ("SEL-021", "Platform", "Global", "Reset water modal background dimmer", "Open hydration reset modal", "Overlay dim background locks screen scroll focus"),
        ("SEL-022", "Groups", "Detail", "Roster podium layout top-3 rankings", "Open leaderboard tab in group", "First place podium tall, second/third adjusted"),
        ("SEL-023", "Groups", "Detail", "Group card layout flex sizing", "Browse group directory cards", "Flex layout wraps goals, codes, member counts"),
        ("SEL-024", "Groups", "Detail", "Profile initials fallback layout", "Load page with name initials avatar", "Renders initials centered inside colored circle"),
        ("SEL-025", "Platform", "Global", "Success toast layout slide-up animations", "Click +250ml water intake", "Mint green toast slides up from bottom, disappears"),
        ("SEL-026", "Platform", "Global", "Checkboxes colors match primary accent", "Toggle reminders setting", "Active check matches primary purple style"),
        ("SEL-027", "Nutrition", "Dashboard", "Hydration ML quick addition transitions", "Click +250ml water button", "Gauge increments immediately without delay"),
        ("SEL-028", "Groups", "Detail", "Leaderboard table scroll bounds", "Populate 25 members to group roster", "Leaderboard table turns scrollable; page stays fixed"),
        ("SEL-029", "Dashboard", "Dashboard", "Macro charts color coordination", "View macro circle indicator chart", "Proteins = red, Carbs = purple, Fats = yellow"),
        ("SEL-030", "Nutrition", "Dashboard", "Food scanning laser alignment guide", "Open food scanner web guide", "Green laser overlay line guide blinks on scanner"),
        ("SEL-031", "Groups", "Detail", "Create Challenge preview matches fields", "Edit target steps in forms", "Live preview card values update matching input"),
        ("SEL-032", "Dashboard", "Dashboard", "Dynamic status indicators connection", "View sync status indicator", "Connection icon turns green when websocket online"),
        ("SEL-033", "Platform", "Global", "Modal barrier click dismissal", "Open hydration reset dialog, click barrier", "Dialog closes safely without committing updates"),
        ("SEL-034", "Auth", "Login", "Sign Up link navigates to register routing", "Click Sign Up navigation link", "Renders registration routing views cleanly"),
        ("SEL-035", "Auth", "Login", "Block URL dashboard navigation if logged out", "Navigate directly to /dashboard", "Route guards clear token, forces back to /login"),
        ("SEL-036", "Platform", "Global", "Axios authorization Bearer header inject", "Inspect profile API request", "Authorization header contains saved JWT token"),
        ("SEL-037", "Platform", "Global", "Clear token headers on logout requests", "Log out, view network calls", "Requests exclude Authorization properties"),
        ("SEL-038", "Platform", "Global", "Axios catch 401 redirect trigger", "Spoof 401 code on profile API", "Clears token local caches, forces /login"),
        ("SEL-039", "Platform", "Global", "Axios timeout alerts under latency", "Simulate network drop on save action", "Axios triggers timeout, alerts user: 'Request timeout'"),
        ("SEL-040", "Platform", "Global", "Double submission click prevention", "Click Save button twice rapidly", "Locks subsequent submit triggers to prevent double hits"),
        ("SEL-041", "Groups", "Detail", "Create group community validation", "Click create group with empty fields", "Blocks submit; prompts: 'Group name is required'"),
        ("SEL-042", "Groups", "Detail", "Invite code characters count limits", "Join group, input 3 character code", "Blocks request; requires 6 alphanumeric characters"),
        ("SEL-043", "Groups", "Detail", "Non-existent group code warning", "Submit invite code 'XXXXXX'", "Warning toast pops: 'Invite code not found'"),
        ("SEL-044", "Groups", "Detail", "Group challenges start dates rules", "Open challenges setup, pick yesterday", "Date validation blocks; requires dates after today"),
        ("SEL-045", "Groups", "Detail", "Challenge end date rules validation", "Set challenge end date before start", "Date picker invalidates, displays format warnings"),
        ("SEL-046", "Groups", "Detail", "Target steps integer validations", "Set steps target to -100 steps", "Blocks; target step field restricts to positive numbers"),
        ("SEL-047", "Profile", "Profile", "Weight inputs boundaries validation", "Input weight of 15kg in profile", "Validation limits trigger: 'Weight must be 20-500kg'"),
        ("SEL-048", "Profile", "Profile", "Height inputs boundaries validation", "Input height of 400cm in profile", "Validation limits trigger: 'Height must be 50-300cm'"),
        ("SEL-049", "Profile", "Profile", "Age inputs boundaries validation", "Input age of 160 in profile settings", "Validation limits trigger: 'Age must be 5-120 years'"),
        ("SEL-050", "Profile", "Profile", "BMR values default calculation safeguards", "Remove weight/height, save profile", "Default state resolves BMR to 1800 kcal safely"),
        ("SEL-051", "Groups", "Detail", "Live chat scroll-to-bottom on click", "Open Chat, scroll up, click 'New Messages'", "Scrolls chat container back to latest message bottom"),
        ("SEL-052", "Groups", "Detail", "Websocket client reconnection behaviors", "Disconnect socket, check network logs", "Client attempts reconnecting automatically every 5s"),
        ("SEL-053", "Platform", "Global", "Network status offline banner layout", "Trigger offline mode on browser", "Red banner 'You are currently offline' renders at top"),
        ("SEL-054", "Nutrition", "Dashboard", "Hydration logs local timezone reset", "Change device date to next day, reload", "Water tracking resets daily counts back to 0ml"),
        ("SEL-055", "Dashboard", "Dashboard", "Zero values empty chart rendering", "Load app with no step history logged", "Web charts render flat 0 heights without crashes"),
        ("SEL-056", "Groups", "Detail", "Leaderboard ranking percentage metrics", "Open challenge progress details", "Displays member percent of goal completed columns"),
        ("SEL-057", "Chat", "Chat", "Blank chat messages blocked", "Type empty space, click send icon", "Send action bypassed; text focus remains active"),
        ("SEL-058", "Platform", "Global", "XSS script inputs sanitization", "Input <script>alert(1)</script> in inputs", "Sanitized on submit, strips tag tags automatically"),
        ("SEL-059", "Platform", "Global", "Upload profile image size limits", "Upload 10MB JPEG image as profile photo", "Warning banner: 'File size exceeds maximum limit of 5MB'"),
        ("SEL-060", "Platform", "Global", "Tab navigation persistent scroll state", "Scroll down dashboard, tap Chat, tap Home", "Dashboard retains original vertical scroll offset"),
        ("SEL-061", "Platform", "Global", "Browser cookie validation configuration", "Inspect cookies under dashboard", "Session cookies set secure flags if HTTPS configured"),
        ("SEL-062", "Platform", "Global", "Multi-tab dashboard sync behaviors", "Open two dashboard tabs, log water in Tab A", "Tab B increments water count automatically"),
        ("SEL-063", "Dashboard", "Dashboard", "Dashboard summary widget loading skeletons", "Open dashboard", "Skeleton card guides render before data populates"),
        ("SEL-064", "Platform", "Global", "Input fields placeholder text visibility", "Inspect email and password input fields", "Placeholder texts are visible with readable color ratio"),
        ("SEL-065", "Auth", "Login", "Login card animation on error", "Submit wrong login credentials", "Card shakes slightly to provide visual error feedback"),
        ("SEL-066", "Auth", "Register", "Password strength indicator bar", "Type 'abc' in password input", "Indicator bar turns red (Weak password feedback)"),
        ("SEL-067", "Auth", "Register", "Password strength medium indicator", "Type 'abc123' in password input", "Indicator bar turns yellow (Medium password feedback)"),
        ("SEL-068", "Auth", "Register", "Password strength strong indicator", "Type 'Abc@12345' in password input", "Indicator bar turns green (Strong password feedback)"),
        ("SEL-069", "Dashboard", "Dashboard", "Toggle calorie goal modal", "Click edit calorie goal target", "Goal modal opens, backdrop dimmed, focus locked"),
        ("SEL-070", "Dashboard", "Dashboard", "Save calorie target edits validation", "Enter calorie goal 50000 kcal", "Error shown: 'Calorie goal must be under 10000 kcal'"),
        ("SEL-071", "Dashboard", "Dashboard", "Calorie goal input rejects letters", "Enter 'abc' in calorie target field", "Field input filters out non-numeric characters"),
        ("SEL-072", "Dashboard", "Dashboard", "Sync progress spinner visibility", "Click Sync on dashboard header", "Spinner icon rotates during API request runtime"),
        ("SEL-073", "Dashboard", "Dashboard", "Sync failure warning alert layout", "Force sync fail on API request", "Red alert banner overlays header: 'Sync failed'"),
        ("SEL-074", "Dashboard", "Dashboard", "Step logs bar charts hover tooltips", "Hover over a column in steps chart", "Tooltip box overlay displays date and exact step counts"),
        ("SEL-075", "Dashboard", "Dashboard", "Calorie surplus pointer coordinates check", "Calorie surplus reads 0 kcal", "Needle pointer remains aligned to exact center line"),
        ("SEL-076", "Nutrition", "Dashboard", "Quick add water buttons hover styles", "Hover mouse over +250ml water button", "Button shifts opacity and scales up slightly (1.05x)"),
        ("SEL-077", "Nutrition", "Dashboard", "Water history weekly charts render", "Open hydration weekly details panel", "Weekly water logs load successfully into bar charts"),
        ("SEL-078", "Nutrition", "Dashboard", "Custom water volume intake input", "Click custom volume input, type '350', add", "Consumed hydration increases by exactly 350ml"),
        ("SEL-079", "Nutrition", "Dashboard", "Custom water intake negative block", "Enter '-200' in custom water volume", "Volume input rejects negative numbers"),
        ("SEL-080", "Nutrition", "Dashboard", "Food search list paginator controls", "Search food oats, scroll to list bottom", "Paginator displays page numbers; loads next results"),
        ("SEL-081", "Nutrition", "Dashboard", "Product nutritional details modal overlay", "Click on oats product card", "Details card opens showing carbohydrates, protein, fats"),
        ("SEL-082", "Nutrition", "Dashboard", "Add calories updates dashboard state", "Add 400 kcal from oats detail panel", "Consumed calorie gauge adds 400 kcal in real-time"),
        ("SEL-083", "Groups", "Detail", "Clipboard copy invite code button hover", "Hover over invite copy button", "Tooltip overlays text: 'Copy to clipboard'"),
        ("SEL-084", "Groups", "Detail", "Clipboard copy success feedback toast", "Click copy invite code button", "Tooltip switches text to 'Copied!'; toast slides up"),
        ("SEL-085", "Groups", "Detail", "Active challenges list timeline styling", "Open group challenges list", "Challenges load in chronological grid timeline layout"),
        ("SEL-086", "Groups", "Detail", "Join challenge member participation", "Click Join Challenge on active list", "User avatar is added to the challenge participants list"),
        ("SEL-087", "Groups", "Detail", "Leave challenge member removal", "Click Leave Challenge on joined challenge", "User avatar removed; steps details updated"),
        ("SEL-088", "Groups", "Detail", "Challenge completion badge overlays", "View completed challenge list", "Finished challenges display gold medal status badges"),
        ("SEL-089", "Chat", "Chat", "Chat messages auto-scroll on send", "Send chat message", "Scrollbar snaps to bottom instantly to show own message"),
        ("SEL-090", "Chat", "Chat", "Chat typing indicator websocket feedback", "Member B starts typing in chat text box", "Message bar displays: 'Member B is typing...'"),
        ("SEL-091", "Chat", "Chat", "Emoticons popup selection window", "Click smiley icon inside chat bar", "Icon tray popup loads showing categories of emoticons"),
        ("SEL-092", "Profile", "Profile", "Toggle system theme dark mode configuration", "Click dark mode toggle in navbar", "Page styling swaps to dark slate theme backgrounds"),
        ("SEL-093", "Profile", "Profile", "Theme preferences persistent storage", "Reload dashboard after setting dark mode", "Theme remains dark; loads state key from localStorage"),
        ("SEL-094", "Profile", "Profile", "Profile height input values format", "Type '180.50' height in profile settings", "Saves decimal values accurately to database profile"),
        ("SEL-095", "Profile", "Profile", "Profile weight input values format", "Type '75.25' weight in profile settings", "Saves decimal values accurately to database profile"),
        ("SEL-096", "Platform", "Global", "External links security target blank checks", "Click Help/Resources link in footer", "Opens help site in new tab using rel=\"noopener noreferrer\""),
        ("SEL-097", "Platform", "Global", "Focus outline keyboard navigation accessibility", "Press Tab key repeatedly to navigate", "Renders clear outline ring focus around links/buttons"),
        ("SEL-098", "Platform", "Global", "Browser window resize layout stability", "Maximize/restore browser window size", "Grid items and charts adapt widths immediately"),
        ("SEL-099", "Platform", "Global", "Browser tab closure session retention", "Close browser window, reopen dashboard URL", "App restores user session if localStorage token valid"),
        ("SEL-100", "Platform", "Global", "Session logout route blocking redirect", "Click Logout, then press browser back button", "Page does not load dashboard; redirects back to login")
    ]
    populate_data_sheet("Selenium Web", selenium_headers, selenium_data)

    # ─── 4. REST & WEBSOCKET API DATA (100 Cases) ──────────────────────────
    api_headers = ["ID", "Module", "Endpoint", "Scenario / Description", "Payload & Query Inputs", "Expected Result"]
    api_data = [
        ("API-001", "Auth", "POST /api/auth/register", "Register new user account successfully", "JSON body containing username, email, password", "Status 200 OK; returns success payload"),
        ("API-002", "Auth", "POST /api/auth/register", "Reject duplicate username registration", "JSON body with existing username", "Status 400 or 500; returns conflict error"),
        ("API-003", "Auth", "POST /api/auth/register", "Reject duplicate email registration", "JSON body with existing email address", "Status 400 or 500; returns conflict error"),
        ("API-004", "Auth", "POST /api/auth/register", "Block blank fields inputs validation check", "JSON with empty fields values", "Status 400 Bad Request; list of error fields"),
        ("API-005", "Auth", "POST /api/auth/register", "Enforce password minimum lengths rules", "Password shorter than 6 characters", "Status 400 Bad Request; length error message"),
        ("API-006", "Auth", "POST /api/auth/register", "Enforce valid email format boundaries", "Malformed email string payload", "Status 400 Bad Request; invalid format error"),
        ("API-007", "Auth", "POST /api/auth/login", "Login with valid credentials check", "JSON containing registered email & password", "Status 200 OK; returns JSON with token key"),
        ("API-008", "Auth", "POST /api/auth/login", "Login with incorrect password validation", "Valid email, wrong password payload", "Status 401 Unauthorized or 403 Forbidden"),
        ("API-009", "Auth", "POST /api/auth/login", "Login with non-existent email verification", "Email not in DB, dummy password payload", "Status 401 Unauthorized or 403 Forbidden"),
        ("API-010", "Profile", "GET /api/profile", "Retrieve profile metadata successfully", "HTTP GET with valid JWT Authorization header", "Status 200 OK; returns profile details JSON"),
        ("API-011", "Profile", "GET /api/profile", "Reject profile queries without JWT header", "HTTP GET, empty Authorization header", "Status 401 Unauthorized; clears routing"),
        ("API-012", "Profile", "PUT /api/profile", "Update profile information successfully", "JSON with updated age, height, weight, gender", "Status 200 OK; returns updated profile JSON"),
        ("API-013", "Profile", "PUT /api/profile", "Validate profile weight minimum boundaries", "JSON weight value under 20kg", "Status 400 Bad Request; weight range error"),
        ("API-014", "Profile", "PUT /api/profile", "Validate profile weight maximum boundaries", "JSON weight value over 500kg", "Status 400 Bad Request; weight range error"),
        ("API-015", "Profile", "PUT /api/profile", "Validate profile height minimum boundaries", "JSON height value under 50cm", "Status 400 Bad Request; height range error"),
        ("API-016", "Profile", "PUT /api/profile", "Validate profile height maximum boundaries", "JSON height value over 300cm", "Status 400 Bad Request; height range error"),
        ("API-017", "Profile", "PUT /api/profile", "Validate profile age minimum boundaries", "JSON age value under 5 years", "Status 400 Bad Request; age range error"),
        ("API-018", "Profile", "PUT /api/profile", "Validate profile age maximum boundaries", "JSON age value over 120 years", "Status 400 Bad Request; age range error"),
        ("API-019", "Profile", "PUT /api/profile", "Enforce gender enum format validations", "Invalid gender string payload (e.g., 'X')", "Status 400 Bad Request or deserialization error"),
        ("API-020", "Profile", "PUT /api/profile", "Reject updates with negative calorie targets", "Daily calorie goal set to -2000 kcal", "Status 400 Bad Request; target range error"),
        ("API-021", "Steps", "POST /api/steps/sync", "Sync user steps log records successfully", "JSON array containing daily dates and steps counts", "Status 200 OK; saves step counts history"),
        ("API-022", "Steps", "POST /api/steps/sync", "Enforce date format YYYY-MM-DD schema", "Date string containing invalid formatting", "Status 400 Bad Request or parsing exception"),
        ("API-023", "Steps", "GET /api/steps/weekly", "Fetch weekly step history logs checklist", "HTTP GET with valid JWT header", "Status 200 OK; returns daily logs array"),
        ("API-024", "Steps", "POST /api/steps/sync", "Update steps counts values for existing date", "POST steps value changes for today's date key", "Status 200 OK; updates existing records in DB"),
        ("API-025", "Steps", "POST /api/steps/sync", "Reject negative step entries on endpoints", "Steps set to -500 in payload array", "Status 400 Bad Request; step range error"),
        ("API-026", "Groups", "POST /api/groups", "Create new fitness group community", "JSON body containing group name & stepGoal", "Status 200 OK; returns group ID & inviteCode"),
        ("API-027", "Groups", "POST /api/groups", "Reject duplicate group names criteria", "JSON containing existing group name", "Status 400 Bad Request or 500 error"),
        ("API-028", "Groups", "GET /api/groups/my", "Retrieve groups user belongs to list", "HTTP GET with valid JWT header", "Status 200 OK; returns list of groups user joined"),
        ("API-029", "Groups", "GET /api/groups/{id}", "Fetch specific group details payload", "Path parameter group ID", "Status 200 OK; returns name, goal, member counts"),
        ("API-030", "Groups", "POST /api/groups/join", "Join group community using inviteCode", "Query parameter code=AB12CD", "Status 200 OK; updates memberships tables"),
        ("API-031", "Groups", "POST /api/groups/join", "Reject joining duplicate group memberships", "Query parameter matching joined group code", "Status 400 Bad Request; duplicate join warning"),
        ("API-032", "Groups", "POST /api/groups/join", "Reject joining with invalid group code", "Query parameter code=XXXXXX", "Status 404 Not Found; error description JSON"),
        ("API-033", "Groups", "GET /api/groups/{id}/leaderboard", "Fetch group step leaderboard rankings", "Path parameter group ID", "Status 200 OK; returns descending members roster"),
        ("API-034", "Groups", "GET /api/groups/{id}/chat/history", "Retrieve group chat history records", "Path parameter group ID", "Status 200 OK; returns JSON message details array"),
        ("API-035", "Challenges", "POST /api/challenges", "Create challenge target inside community", "JSON containing name, target steps, dates range", "Status 200 OK; maps challenge to group"),
        ("API-036", "Challenges", "GET /api/challenges/group/{id}", "Fetch active challenges list for group", "Path parameter group ID", "Status 200 OK; returns array of challenges"),
        ("API-037", "Challenges", "GET /api/challenges/{id}/progress", "Fetch challenge leaderboard percentages", "Path parameter challenge ID", "Status 200 OK; returns sorted users progress lists"),
        ("API-038", "Chat", "WebSocket ws-chat", "STOMP client handshake authorization check", "WebSocket handshake, verify JWT header", "Handshake established successfully if token valid"),
        ("API-039", "Chat", "WebSocket ws-chat", "STOMP channel subscription auth guard", "Subscribe path /topic/group.{id}", "Subscription accepted if user is group member"),
        ("API-040", "Chat", "WebSocket ws-chat", "STOMP broadcast text message payload", "Publish message to chat topic destination", "Message broadcasted to all subscribers in group"),
        ("API-041", "Chat", "WebSocket ws-chat", "Reject chat messages with blank content", "Publish message empty body JSON payload", "Message rejected; no broadcast triggered"),
        ("API-042", "Platform", "Security", "SQL Injection input sanitization checks", "Send payload containing SQL query codes", "Inputs treated as literals; query parameters safe"),
        ("API-043", "Platform", "Security", "Cross-site Scripting XSS tags stripping", "Send payload containing script HTML tag fields", "Server escapes or strips HTML tags on database write"),
        ("API-044", "Platform", "Security", "CORS policies config check", "Options preflight call from unauthorized origin", "Status 403 Forbidden; origin blocked"),
        ("API-045", "Platform", "Security", "Content Security Policy headers verify", "HTTP GET request headers response inspection", "Headers check out: CSP, X-Frame-Options, HSTS"),
        ("API-046", "Platform", "Security", "Rate limiting safety controls", "Make 200 calls in under 10 seconds", "Status 429 Too Many Requests; blocks IP temporary"),
        ("API-047", "Platform", "Security", "Database transactions rollback on errors", "Trigger fail step halfway through complex transaction", "DB rollback executes; data remains consistent"),
        ("API-048", "Platform", "Global", "Global Exception Handler formats", "Force validation error on JSON payloads", "Error returns uniform JSON payload structure"),
        ("API-049", "Platform", "Performance", "Endpoint response latency threshold", "Measure latency on /api/steps/weekly", "Response time under 200ms for standard requests"),
        ("API-050", "Platform", "Global", "API log levels levels configuration", "Trigger request execution", "Console prints info logs; debug suppressed in prod"),
        ("API-051", "Auth", "POST /api/auth/register", "Username regex validation checks", "Register username with symbols usr!@#", "Status 400 Bad Request; alphanumeric validation"),
        ("API-052", "Auth", "POST /api/auth/register", "Email max length constraint", "Register email with 300 characters string", "Status 400 Bad Request; length validator error"),
        ("API-053", "Auth", "POST /api/auth/register", "Password strength requirement checks", "Register weak password (e.g., '123')", "Status 400 Bad Request; weak password error"),
        ("API-054", "Auth", "POST /api/auth/login", "Login schema validation checks", "Send login payload with missing email field", "Status 400 Bad Request; validation failed"),
        ("API-055", "Auth", "GET /api/auth/validate-token", "Validate JWT token expiration checks", "HTTP GET with expired JWT header token", "Status 401 Unauthorized; token expired"),
        ("API-056", "Profile", "PUT /api/profile", "Calorie goal maximum boundaries", "Set daily calorie goal to 20,000 kcal", "Status 400 Bad Request; max calorie range error"),
        ("API-057", "Profile", "PUT /api/profile", "Height values negative check", "Set profile height to -180.0 cm", "Status 400 Bad Request; must be positive value"),
        ("API-058", "Profile", "PUT /api/profile", "Weight values negative check", "Set profile weight to -75.0 kg", "Status 400 Bad Request; must be positive value"),
        ("API-059", "Profile", "PUT /api/profile", "Gender field blank error", "Set profile gender to empty string", "Status 400 Bad Request; gender field required"),
        ("API-060", "Profile", "GET /api/profile/achievements", "Fetch unlocked achievements lists", "HTTP GET with valid JWT auth header", "Status 200 OK; returns array of unlocked badges"),
        ("API-061", "Steps", "POST /api/steps/sync", "Step log duplicate date handling", "Post sync step logs twice for date '2026-06-16'", "Status 200 OK; updates step count instead of duplicate"),
        ("API-062", "Steps", "POST /api/steps/sync", "Calories values negative block", "Post steps sync calories count to -300 kcal", "Status 400 Bad Request; calorie range error"),
        ("API-063", "Steps", "POST /api/steps/sync", "Distance values negative block", "Post steps sync distance count to -5.0 km", "Status 400 Bad Request; distance range error"),
        ("API-064", "Steps", "GET /api/steps/monthly", "Retrieve monthly steps summary data", "HTTP GET with valid JWT auth header", "Status 200 OK; returns monthly summary data"),
        ("API-065", "Steps", "GET /api/steps/yearly", "Retrieve yearly steps summary data", "HTTP GET with valid JWT auth header", "Status 200 OK; returns yearly summary data"),
        ("API-066", "Groups", "POST /api/groups", "Group name minimum length validation", "Create group with name 'A'", "Status 400 Bad Request; name too short error"),
        ("API-067", "Groups", "POST /api/groups", "Group name maximum length validation", "Create group with name 100 chars long", "Status 400 Bad Request; name too long error"),
        ("API-068", "Groups", "DELETE /api/groups/{id}", "Delete group community validation", "Delete group request by group ID as admin", "Status 200 OK; group removed, members notified"),
        ("API-069", "Groups", "DELETE /api/groups/{id}", "Non-admin user delete group block", "Delete group request by group ID as regular member", "Status 403 Forbidden; authorization denied"),
        ("API-070", "Groups", "POST /api/groups/{id}/leave", "Leave group community endpoint", "POST request to leave group by group ID", "Status 200 OK; member removed from group database"),
        ("API-071", "Groups", "POST /api/groups/{id}/kick/{userId}", "Kick member from group as admin", "POST kick user request by group ID and user ID", "Status 200 OK; user kicked from community DB"),
        ("API-072", "Groups", "POST /api/groups/{id}/kick/{userId}", "Non-admin kick member execution block", "POST kick user request as regular member", "Status 403 Forbidden; action denied"),
        ("API-073", "Groups", "GET /api/groups/search", "Search groups by keyword search query", "HTTP GET query ?q=Team", "Status 200 OK; returns matched groups list"),
        ("API-074", "Challenges", "POST /api/challenges", "Challenge name missing validation", "Create challenge with empty name field", "Status 400 Bad Request; name required error"),
        ("API-075", "Challenges", "POST /api/challenges", "Challenge step goal under limits check", "Create challenge with target steps = 50 steps", "Status 400 Bad Request; step goal too low"),
        ("API-076", "Challenges", "POST /api/challenges", "Challenge dates overlap validation", "Set challenge start date after end date", "Status 400 Bad Request; start date after end"),
        ("API-077", "Challenges", "GET /api/challenges/active", "Retrieve global user active challenges list", "HTTP GET with valid JWT auth header", "Status 200 OK; returns array of user challenges"),
        ("API-078", "Challenges", "POST /api/challenges/{id}/join", "Join active group challenge community", "POST join request by challenge ID", "Status 200 OK; maps member to challenge progress"),
        ("API-079", "Challenges", "POST /api/challenges/{id}/join", "Duplicate challenge join validation", "POST join request twice for same challenge", "Status 400 Bad Request; already participating"),
        ("API-080", "Challenges", "POST /api/challenges/{id}/leave", "Leave challenge community validation", "POST leave request by challenge ID", "Status 200 OK; user progress records cleared"),
        ("API-081", "Chat", "WebSocket ws-chat", "Send message to non-existing group topic", "Publish message to /topic/group.9999 (invalid ID)", "Subscription error; broker closes communication"),
        ("API-082", "Chat", "WebSocket ws-chat", "STOMP chat message length bounds validation", "Send message body containing 2000 characters", "Payload size exception; message truncated/rejected"),
        ("API-083", "Chat", "WebSocket ws-chat", "Send media attachments metadata payload", "Publish message with image url attachment link", "Status 200 OK; message metadata saves with link"),
        ("API-084", "Chat", "WebSocket ws-chat", "STOMP unsubscribe channel routing", "Unsubscribe from group chat topic", "Broker closes subscription, user no longer receives"),
        ("API-085", "Chat", "WebSocket ws-chat", "WebSocket socket session disconnection", "Close active websocket client connection", "Broker registers close session, marks user offline"),
        ("API-086", "Platform", "Security", "SQL wildcards escape validation in search", "Send query ?q=% on group search", "Status 200 OK; query returns matched text properly"),
        ("API-087", "Platform", "Security", "Path Traversal request parameter validations", "Query /api/profile?file=../../etc/passwd", "Status 400 Bad Request or parameters ignored"),
        ("API-088", "Platform", "Security", "JWT Signature verification safety checks", "Send API request with modified JWT token key", "Status 401 Unauthorized; signature validation fail"),
        ("API-089", "Platform", "Security", "CORS origin allowlist checks validation", "Request API from allowed origin (localhost:5173)", "Status 200 OK; contains Access-Control-Allow-Origin"),
        ("API-090", "Platform", "Performance", "Concurrent steps sync stress test", "Trigger 50 concurrent sync POST requests", "Status 200 OK; database locks managed, no failures"),
        ("API-091", "Platform", "Performance", "Leaderboard aggregation query speed", "Fetch leaderboard for group with 100 members", "Response resolved under 100ms; index optimized"),
        ("API-092", "Platform", "Global", "Content type headers JSON compliance", "Send GET request to /api/profile", "Response header contains Content-Type: application/json"),
        ("API-093", "Platform", "Global", "Request compression header compression", "Send GET with Accept-Encoding: gzip", "Server responds with gzip compression headers"),
        ("API-094", "Platform", "Global", "PUT profile with decimal values", "Update weight to 75.35 kg, height to 180.25 cm", "Status 200 OK; floats saved correctly in database"),
        ("API-095", "Platform", "Global", "Invalid JSON payload format handler", "Send malformed JSON payload structure", "Status 400 Bad Request; malformed JSON error"),
        ("API-096", "Platform", "Security", "Secure WebSocket (WSS) protocol verify", "Connect to ws-chat endpoint", "Handshake completes on wss:// protocol (if prod)"),
        ("API-097", "Platform", "Performance", "Large message history fetch payload speed", "Fetch history containing 500 chat messages", "Status 200 OK; response compressed under 150ms"),
        ("API-098", "Platform", "Security", "JWT Authorization token header prefixes", "Send token without Bearer prefix keyword", "Status 401 Unauthorized; validation fails"),
        ("API-099", "Platform", "Security", "Content Security Policy Frame options", "Inspect HTTP headers", "Response contains X-Frame-Options: DENY"),
        ("API-100", "Platform", "Global", "REST API Healthcheck status endpoint", "HTTP GET request to /actuator/health or /health", "Status 200 OK; returns {\"status\":\"UP\"}")
    ]
    populate_data_sheet("REST & WebSocket API", api_headers, api_data)

    output_path = r"d:\App\CircleFit\CircleFit_Master_Test_Cases_300.xlsx"
    wb.save(output_path)
    print(f"[OK] Master excel sheet saved successfully at: {output_path}")

if __name__ == "__main__":
    main()
